"""Excel ingestion for the MoDoc Q&A content source."""

from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class QnaSource:
    """Normalized source record used by the generation pipeline."""

    row_number: int
    question_text: str
    expert_answer_text: str
    status_english: str
    english_date: str
    wix_post_url_english: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def load_qna_sources(
    input_path: Path,
    *,
    status_filter: str = "Published",
    row_number: int | None = None,
    limit: int = 1,
) -> list[QnaSource]:
    """Load Q&A rows from the workbook.

    Published rows are the default because the internship pipeline should
    convert content that already passed the blog publishing workflow. The
    caller can override this with `--status any` or `--row` when testing.
    """

    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    header_index = {header: index for index, header in enumerate(headers) if header}

    required_headers = ["Question text", "Expert Answer text", "Status (English)"]
    missing = [header for header in required_headers if header not in header_index]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}")

    selected: list[QnaSource] = []
    for excel_row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if row_number is not None and excel_row_number != row_number:
            continue

        source = _row_to_source(row, header_index, excel_row_number)
        if not source.question_text or not source.expert_answer_text:
            continue

        if row_number is None and not _matches_status(source.status_english, status_filter):
            continue

        selected.append(source)
        if len(selected) >= limit:
            break

    if not selected:
        scope = f"row {row_number}" if row_number is not None else f"status '{status_filter}'"
        raise ValueError(f"No usable Q&A rows found for {scope}.")

    return selected


def _row_to_source(
    row: tuple[Any, ...],
    header_index: dict[str, int],
    excel_row_number: int,
) -> QnaSource:
    return QnaSource(
        row_number=excel_row_number,
        question_text=_cell_text(row, header_index, "Question text"),
        expert_answer_text=_cell_text(row, header_index, "Expert Answer text"),
        status_english=_cell_text(row, header_index, "Status (English)"),
        english_date=_cell_text(row, header_index, "English date"),
        wix_post_url_english=_cell_text(row, header_index, "WIX Post url (English)"),
    )


def _cell_text(row: tuple[Any, ...], header_index: dict[str, int], header: str) -> str:
    index = header_index.get(header)
    if index is None or index >= len(row):
        return ""
    return _stringify_cell(row[index])


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _matches_status(actual: str, expected: str) -> bool:
    if expected.lower() == "any":
        return True
    return actual.strip().lower() == expected.strip().lower()

